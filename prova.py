import pandas as pd
import streamlit as st
import mysql.connector
import io
import mysql.connector
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import locale
locale.setlocale(locale.LC_ALL, 'es_ES.utf8')

mydb = mysql.connector.connect(
  host="109.232.71.246",
  user="prova",
  password="123456789",
  database="esp_data"
)

mycursor = mydb.cursor()

mycursor.execute(f"SELECT * FROM regis")
myresult = mycursor.fetchall()  
test = pd.DataFrame(myresult) 
test.columns = ["id","nombre","nfcid","accion","lugar","Fecha_Hora"]


# borrar datos sin importancia del pagina como iconitas con css, la clase se encuentra con click derecha y inspectionar
st.markdown("""
<style>
.css-nqowgj.e1ewe7hr3{
    visibility: hidden;
}

.css-h5rgaw.e1g8pov61{
    visibility: hidden;
}
.css-zt5igj.eqr7zpz3 {
    color: darkslategray;
    text-align: center;
    box-shadow: rgba(0, 0, 0, 0.45) 0px 25px 20px -20px;
    padding: 1rem;
body {
    background-color: darkslategray;
}
.css-1kyxreq.ebxwdo62 {
    height: 3px;
    width: 3px;
    border-radius: 50px;
    text-align: center;
}
</style>
""", unsafe_allow_html=True)

# Titulo
st.title('Resultados por dia:')

# colorear la fila
def apply_row_colors(row):
    if row['lugar'] == 'SALIDA FORZADA':
        return ['background-color: red'] * len(row)
    else:
        return [''] * len(row)

test['Fecha_Hora'] = pd.to_datetime(test['Fecha_Hora'])

selected_date = st.date_input("Select a date")
selected_nombre = st.selectbox("Select a name", test['nombre'].unique())


if selected_date:
    filtered_data = test[(test['Fecha_Hora'].dt.date == selected_date) & (test['nombre'] == selected_nombre)]
    if not filtered_data.empty:
        entrada_indices = filtered_data[filtered_data['accion'] == 'Entrada'].index
        salida_indices = filtered_data[filtered_data['accion'] == 'Salida'].index
        
        total_duration = pd.Timedelta(0)
        
        for entrada_index in entrada_indices:
            try:
                salida_index = salida_indices[salida_indices > entrada_index][0]
                duration = filtered_data.loc[salida_index, 'Fecha_Hora'] - filtered_data.loc[entrada_index, 'Fecha_Hora']
                total_duration += duration
            except IndexError:
                continue
        
        total_duration_hours = total_duration.total_seconds() // 3600
        total_duration_minutes = (total_duration.total_seconds() % 3600) // 60
        total_duration_seconds = total_duration.total_seconds() % 60

        #----------------------------------------------------------------------------------------------------------------------
        entrada_df = filtered_data[filtered_data['accion'] == 'Entrada'][['id', 'nombre', 'nfcid', 'lugar', 'Fecha_Hora']].rename(columns={'Fecha_Hora': 'entrada'})
        salida_df = filtered_data[filtered_data['accion'] == 'Salida'][['id', 'nombre', 'nfcid', 'lugar', 'Fecha_Hora']].rename(columns={'Fecha_Hora': 'salida'})

        # Merge 'entrada' and 'salida' DataFrames on common columns
        merged_df = pd.merge(entrada_df, salida_df, on=['id', 'nombre', 'nfcid', 'lugar'], how='outer')

        # Sort the DataFrame by 'id' and 'entrada' date
        merged_df.sort_values(by=['id', 'entrada'], inplace=True)

        # Reset the index
        merged_df.reset_index(drop=True, inplace=True)

        # Shift 'entrada' values down by one row
        merged_df['entrada'] = merged_df['entrada'].shift(1)
        #Delete the row without values
        merged_df.dropna(subset=['salida'], inplace=True)

  
        

        merged_df['Date'] = pd.to_datetime(merged_df["entrada"])
        merged_df['Date'] = pd.to_datetime(merged_df['Date']).dt.date

        # Set the 'date' column as the DataFrame index
        merged_df = merged_df.set_index('Date')
        merged_df['entrada'] = pd.to_datetime(merged_df['entrada'], format='%H:%M').dt.time
        merged_df['salida'] = pd.to_datetime(merged_df['salida'], format='%H:%M').dt.time

        
        # Apply row colors using the apply() method with axis=1
        merged_df = merged_df.reset_index(drop=True)
        styled_data = merged_df[["id", "nombre", "nfcid", "lugar", "entrada", "salida"]].style.apply(apply_row_colors, axis=1)

        st.write(styled_data, hide_index=True, width=1000)
            
        st.write("Total Time Duration:")
        st.write("Hours:", total_duration_hours)
        st.write("Minutes:", total_duration_minutes)
        st.write("Seconds:", total_duration_seconds)
    else:
        st.write("No records found for the selected date and name.")



st.markdown("---")
st.title('Resultados por mes:')


selected_month = st.selectbox("Select a month", range(1,13))
selected_year = st.selectbox("Select a year", test['Fecha_Hora'].dt.year.unique())
selected = st.selectbox("Elige nombre: ", test['nombre'].unique())



if selected_month and selected_year and selected:
    start_date = pd.to_datetime(f"{selected_year}-{selected_month}-01")
    end_date = start_date + pd.offsets.MonthEnd(1)
    filtered_data = test[(test['Fecha_Hora'] >= start_date) & (test['Fecha_Hora'] <= end_date) & (test['nombre'] == selected)]
else:
    filtered_data = pd.DataFrame()

if not filtered_data.empty:
    entrada_indices = filtered_data[filtered_data['accion'] == 'Entrada'].index
    salida_indices = filtered_data[filtered_data['accion'] == 'Salida'].index
    
    total_duration = pd.Timedelta(0)
    
    for entrada_index in entrada_indices:
        try:
            salida_index = salida_indices[salida_indices > entrada_index][0]
            duration = filtered_data.loc[salida_index, 'Fecha_Hora'] - filtered_data.loc[entrada_index, 'Fecha_Hora']
            total_duration += duration
        except IndexError:
            continue
    
    total_duration_hours = total_duration.total_seconds() // 3600
    total_duration_minutes = (total_duration.total_seconds() % 3600) // 60
    total_duration_seconds = total_duration.total_seconds() % 60

       #----------------------------------------------------------------------------------------------------------------------
    entrada_df = filtered_data[filtered_data['accion'] == 'Entrada'][['id', 'nombre', 'nfcid', 'lugar', 'Fecha_Hora']].rename(columns={'Fecha_Hora': 'entrada'})
    salida_df = filtered_data[filtered_data['accion'] == 'Salida'][['id', 'nombre', 'nfcid', 'lugar', 'Fecha_Hora']].rename(columns={'Fecha_Hora': 'salida'})

    # Merge 'entrada' and 'salida' DataFrames on common columns
    merged_df = pd.merge(entrada_df, salida_df, on=['id', 'nombre', 'nfcid', 'lugar'], how='outer')

    # Sort the DataFrame by 'id' and 'entrada' date
    merged_df.sort_values(by=['id', 'entrada'], inplace=True)

    # Reset the index
    merged_df.reset_index(drop=True, inplace=True)

    # Shift 'entrada' values down by one row
    merged_df['entrada'] = merged_df['entrada'].shift(1)
    #Delete the row without values
    merged_df.dropna(subset=['salida'], inplace=True)

    merged_df["Fecha_Hora"] = pd.to_datetime(merged_df["entrada"])
    

    merged_df['Date'] = merged_df['Fecha_Hora'].dt.date
    merged_df['Date'] = pd.to_datetime(merged_df['Date'])

    # Set the 'date' column as the DataFrame index
    merged_df = merged_df.set_index('Date')

    month = merged_df.index.month[0]
    year = merged_df.index.year[0]

    # Define the date range for the month
    start_date = pd.to_datetime(f'{year}-{month:02d}-01')
    end_date = start_date + pd.offsets.MonthEnd()
    date_range = pd.date_range(start=start_date, end=end_date, freq='D')

        # Create a new DataFrame with the complete date range
    complete_dates = pd.DataFrame(index=date_range)
    
    # Merge the existing DataFrame with the complete date DataFrame
    
    data = complete_dates.merge(merged_df, how='left', left_index=True, right_index=True)

    # Reset the index to make the 'date' column a regular column
    data = data.reset_index()

    
    data["diaSemana"] = pd.to_datetime(data["index"]).dt.day_name(locale='es_ES.utf8')

    data['Tot'] = data['salida'] - data['entrada']



     # --------------------------------Change the format for some columns-------------------------
    data['entrada'] = pd.to_datetime(data['entrada'], format='%H:%M').dt.time
    data['salida'] = pd.to_datetime(data['salida'], format='%H:%M').dt.time
    data['index'] = data['index'].dt.date
    # ------------------------------------Delete unnecesary columns----------------------------------
    data = data.drop(["id", "Fecha_Hora"], axis=1)

     # ----------------------------- Create column daily_result --------------------
    # Convert the 'index' column to datetime format
    data['index'] = pd.to_datetime(data['index'])
    # Convert the 'Tot' column to timedelta format
    data['Tot'] = pd.to_timedelta(data['Tot'])
    # Group the DataFrame by day and calculate the sum of 'Tot' for each day
    data['daily_result'] = data.groupby(data['index'].dt.date)['Tot'].transform('sum')   


    # Convert the 'daily_result' column to timedelta format
    data['daily_result'] = pd.to_timedelta(data['daily_result'])
    data['Tot'] = np.where(data['entrada'].notnull(), data['Tot'].astype(str).str[-8:], '')
    data['daily_result'] = np.where(data['entrada'].notnull(), data['daily_result'].astype(str).str[-8:], '')

    # ------------------------------------Delete the first result when data is repeating------------------------
    data.loc[data['index'].duplicated(keep='first'), 'daily_result'] = pd.NA

    #-----------------------------------------------------------------------------------
    # ----show streamlit
    data_streamlit_show = data.copy(deep=True)

    data_streamlit_show["index"] = pd.to_datetime(data_streamlit_show["index"]).dt.date

    # Delete column unaccessary
    data_streamlit_show = data_streamlit_show.drop(["nombre", "nfcid", "daily_result"], axis=1)

       # Apply row colors using the apply() method with axis=1
    styled_data = data_streamlit_show[["index", "lugar", "entrada", "salida", "diaSemana", "Tot"]].style.apply(apply_row_colors, axis=1)

    # show dataframe
    st.dataframe(styled_data, hide_index=True, width=800, height=400)

    st.write("Total Time Duration:")
    st.write("Hours:", total_duration_hours)
    st.write("Minutes:", total_duration_minutes)
    st.write("Seconds:", total_duration_seconds)

        #-------------------------------Add column Total----------------------

    data['Total'] = data['daily_result']
    
    data['daily_result'] = pd.to_timedelta(data['daily_result'].fillna('00:00:00'))

    total_hours_sum = data['daily_result'].sum()
    

    sum_hours = total_hours_sum.days * 24 + total_hours_sum.seconds // 3600
    sum_minutes = (total_hours_sum.seconds % 3600) // 60
    sum_seconds = total_hours_sum.seconds % 60

    sum_row = pd.DataFrame({'daily_result': [total_hours_sum]})
    sum_row['Total'] = "Total por mes: {:02d}:{:02d}:{:02d}".format(sum_hours, sum_minutes, sum_seconds)

    

    sum_row_hours = pd.DataFrame({'daily_result': [total_hours_sum]})
    sum_row_hours['Total'] = "Total horas: {:02d}".format(sum_hours)

    sum_row_minutes = pd.DataFrame({'daily_result': [total_hours_sum]})
    sum_row_minutes['Total'] = "Total minutes: {:02d}".format(sum_minutes)

    data = data.append(sum_row, ignore_index=True)
    data = data.append(sum_row_hours, ignore_index=True)
    data = data.append(sum_row_minutes, ignore_index=True)

    # Variables with the data for merge cells with openpysl
    nombre = data.iloc[2][1]
    lugar = data.iloc[2][4]
    nfcid = data.iloc[2][2]
    fecha = data["index"].dt.month_name(locale='es_ES.utf8')

    

        #------------------------------Rename, delete the columns--------------------------------
    data.rename(columns={'index':'Date','nombre': 'Nombre', 'entrada' : 'Entrada', 'salida' : 'Salida'}, inplace=True)

    #data['Data'] = pd.to_datetime(data["index"]).dt.date
    
    
    data = data.drop(["Nombre", "nfcid", "daily_result"], axis=1)

    #data = data.reindex(columns=["Data", "lugar", "diaSemana", "Entrada", "Salida", "Tot", "Total"])
    
    #data.drop(data.columns[3], axis=1, inplace=True)
    #data = data[["Date", "index", "lugar", "Entrada", "Salida", "Tot", "Total"]]

    data['Date'] = data['Date'].dt.date

    

    # ------------------------------------hightlight rows--------------------------
    
    def highlight_days(row):

        styles = [''] * len(row)

        if row['diaSemana'] == 'Sábado' or row['diaSemana'] == 'Domingo':
            styles = ['background-color: palegoldenrod'] * len(row)

        elif row['lugar'] == 'SALIDA FORZADA':
            styles = ['background-color: salmon'] * len(row)


        return styles
    
    data = data.style.apply(lambda row: highlight_days(row), axis=1, subset=pd.IndexSlice[:, :])

    
    data.to_excel(r"C:\\Users\\Andrew\\Desktop\\fibwi\\resuldatos\\df_new.xlsx", index=False)






    #----------------------------------------------------------------------------------------------------------------------

     # ---------------------------------------OPEMPYXL-------------------------------------------------------

    book = load_workbook('df_new.xlsx')

    # Select the sheet in the Excel file
    sheet = book.active
 
    #Delete column Date without name
    #delete_column_date = 1
    #sheet.delete_cols(delete_column_date) 

    sheet.insert_cols(idx=1)



    #Adjust the width of the columns
    sheet.column_dimensions['A'].width = 23
    sheet.column_dimensions['B'].width = 18
    sheet.column_dimensions['C'].width = 18
    sheet.column_dimensions['D'].width = 18
    sheet.column_dimensions['E'].width = 18
    sheet.column_dimensions['F'].width = 18
    sheet.column_dimensions['G'].width = 18
    sheet.column_dimensions['H'].width = 21

    

    # Insert empty rows
    sheet.insert_rows(1, 3)
    # Merge cells
    sheet.merge_cells('B1:H1')
    sheet.merge_cells('B2:H2')
    sheet.merge_cells('B3:H3')

    # Define the current column index and the new column name
    current_column_index = 2  # Replace with the current column index (1-based)
    new_column_name = 'Date'  # Replace with the desired new column name

    # Update the value of the cell in the first row and current column index
    sheet.cell(row=4, column=current_column_index, value=new_column_name)

    

    sheet['B1'] = 'Nombre: ' + str(nombre)
    sheet['B2'] = fecha.iloc[1] 
    sheet['B3'] = 'NFCID: ' + str(nfcid)

    #Style cell nombre
    b1_nombre = sheet['B1']
    ft = Font(color="FF0000")
    b1_nombre.font = ft


# ---------------------------Merge column Date, diaSemana----------------------
    # Specify the column index you want to merge rows in
    column_index = 2  

    # Initialize variables to track the start and end rows of consecutive repeating values
    start_row = 5  # Start from row 2
    prev_value = sheet.cell(row=start_row, column=column_index).value
    

    # Iterate over the cells in the column
    for row in range(start_row + 1, sheet.max_row + 1):
        current_value = sheet.cell(row=row, column=column_index).value
        
        if current_value == prev_value:
            # Continue the merging range
            prev_value = current_value
        else:
            # Merge the cells if the previous value is repeating
            if row - start_row > 1:
                sheet.merge_cells(start_row=start_row, start_column=column_index, end_row=row - 1, end_column=column_index)
                column_index += 4 
                sheet.merge_cells(start_row=start_row, start_column=column_index, end_row=row - 1, end_column=column_index)
                #column_index -= 4
                
                column_index += 2
                sheet.merge_cells(start_row=start_row, start_column=column_index, end_row=row - 1, end_column=column_index)
                column_index -= 6
                
            
            # Update the start row and previous value for the new range
            start_row = row
            prev_value = current_value


    # --------------------------------------Align center-------------------------

    # Define the alignment style
    alignment = Alignment(horizontal='center', vertical='center')

    # Iterate through all cells in the worksheet
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = alignment  # Apply the alignment style



    #.---------------------------------------- color SALIDA FORZADA y cuando hay merge-------------------------


    for row in sheet.iter_rows(min_row=4):
        for cell in row:
            if cell.value == 'SALIDA FORZADA':
                # Apply the fill color to the entire row
                for row_cell in row[1:]:
                    row_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    if row[1].value == sheet[row[0].row - 2][0].value:
                    # Apply the fill color to the row above
                        for row_cell in sheet[row[0].row - 1][1:]:
                            row_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                break


    # Create a download button for the filtered data and total duration
    buffer = io.BytesIO()
    book.save(buffer)
    buffer.seek(0)

    st.download_button(
        label="Download Filtered Data",
        data=buffer,
        file_name="filtered_data.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

else:
    if selected_month and selected_year:
        st.write("No records found for the selected month, year, and name.")
    else:
        st.write("Please select a month, year, and name.")


#val = st.text_area("Details:")





